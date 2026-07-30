#!/usr/bin/env python3
"""Build the automated Associate Allocations workbook.

Reads the legacy hand-maintained workbook (block-per-week layout on
"SOA Allocation 2026" plus a hand-typed round robin on "Peer Review
Allocations") and writes a formula-driven workbook where:

  * the week calendar is generated from one anchor date,
  * allocations live in one flat, filterable register,
  * peer reviewers are assigned by an automatic round robin,
  * load vs capacity is visible per week, per associate and per division.

Usage:
    python build_workbook.py SOURCE.xlsx OUTPUT.xlsx

Only classic (pre-2007) worksheet functions are used, so the file behaves
the same in Excel desktop, Excel for the web and Google Sheets.
"""

from __future__ import annotations

import datetime as dt
import re
import sys
from collections import Counter, OrderedDict

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------
# House style
# --------------------------------------------------------------------------
FONT = "Arial"

H1 = Font(name=FONT, size=14, bold=True, color="FF1F3864")
H2 = Font(name=FONT, size=11, bold=True, color="FF1F3864")
HDR = Font(name=FONT, size=10, bold=True, color="FFFFFFFF")
BODY = Font(name=FONT, size=10)
BODY_B = Font(name=FONT, size=10, bold=True)
CALC = Font(name=FONT, size=10, color="FF1F3864")
MUTED = Font(name=FONT, size=9, italic=True, color="FF7F7F7F")
HELP_HDR = Font(name=FONT, size=9, bold=True, color="FF7F7F7F")

FILL_HDR = PatternFill("solid", fgColor="FF1F3864")      # dark blue header
FILL_SUB = PatternFill("solid", fgColor="FFD9E1F2")      # pale blue sub-header
FILL_IN = PatternFill("solid", fgColor="FFFFF2CC")       # yellow  = type here
FILL_CALC = PatternFill("solid", fgColor="FFF2F2F2")     # grey    = automatic
FILL_HELP = PatternFill("solid", fgColor="FFE7E6E6")     # helper columns
FILL_OK = PatternFill("solid", fgColor="FFE2EFDA")       # green
FILL_WARN = PatternFill("solid", fgColor="FFFFE699")     # amber
FILL_BAD = PatternFill("solid", fgColor="FFFFC7CE")      # red
FILL_NOW = PatternFill("solid", fgColor="FFDDEBF7")      # current week

THIN = Side(style="thin", color="FFBFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_FMT = "dd/mm/yyyy"
LAST_ALLOC_ROW = 900           # allocations register: rows 2..900
N_ASSOC_COLS = 12              # associate columns on the dashboard / week view
N_WEEK_ROWS = 53               # weeks generated

# --------------------------------------------------------------------------
# Reference data
# --------------------------------------------------------------------------
# Short name -> (full name, role, active, SOAs per week, in peer-review rotation)
# Short names are the values already used on the legacy allocation tab; full
# names are the ones used on the legacy peer review tab.
ROSTER = [
    ("Addison", "Addison Nolan", "Associate", "Y", 3, "Y", ""),
    ("Chloe", "Chloe", "Associate", "Y", 3, "N", ""),
    ("Chris H", "Chris H", "Associate", "Y", 3, "N", ""),
    ("Cristina", "Cristina Flores", "Associate", "Y", 3, "Y", ""),
    ("Harry", "Harry Devine", "Associate", "Y", 3, "Y", ""),
    ("Ikesha", "Ikesha King", "Team leader", "Y", 2, "Y", ""),
    ("James", "James", "Associate", "Y", 3, "Y", ""),
    ("Josh Bevan", "Joshua Bevan", "Associate", "Y", 3, "Y", ""),
    ("Kyle", "Kyle Wieser", "Associate", "Y", 3, "Y", ""),
    ("Srushti", "Srushti", "Associate", "Y", 3, "Y", ""),
    ("TP", "TP", "Associate", "Y", 3, "N", ""),
    ("Chris R", "Chris R", "Adviser", "N", 0, "N",
     "Appears once as an associate on the 2026 tab"),
    ("Tyler (IK training)", "Tyler", "Trainee", "N", 0, "N",
     "Legacy value kept so 2026 history still matches the roster"),
    ("Henry Chen", "Henry Chen", "Associate", "N", 3, "N", "Left rotation"),
    ("Hayden Wilson", "Hayden Wilson", "Associate", "N", 3, "N", "Left rotation"),
    ("Lachlan Brien", "Lachlan Brien", "Associate", "N", 3, "N", "Left rotation"),
    ("Shannon Victor", "Shannon Victor", "Associate", "N", 3, "N", "Left rotation"),
]

ADVISERS = ["Andrew", "Ben", "Chris R", "Colby", "Cory", "Damien", "Guy",
            "Hayden B", "Izzy", "Josh Bareham", "Matt", "MWS Life", "Nathan",
            "Sam", "Wilson"]

DIVISIONS = ["MWS1", "MWS2", "MWS3", "EWAR", "MWS Life"]

CLIENT_TYPES = ["New Business", "Existing", "Insurance only", "Amendment"]

PROGRESS = ["Not started", "Data completed", "Drafting", "With peer reviewer",
            "Peer review done", "Sent to adviser", "Completed", "On hold",
            "Cancelled", "Rescheduled"]

APPT_STATUS = ["Booked", "No appt", "TBC", "Not booked yet", "N/A"]

REVIEW_STATUS = ["Not uploaded", "Awaiting review", "In review",
                 "Changes requested", "Approved"]

# Division weekly caps, carried over from the "Assoc. Caps" / "Imp. Caps"
# blocks on the legacy 2025 tab (A8:B18).
DIVISION_CAPS = [("MWS1", 8, 8), ("MWS2", 3, 2), ("MWS3", 3, 2),
                 ("EWAR", 5, 2), ("MWS Life", 2, 2)]

# 2026 public holidays. The national dates plus the NSW/QLD dates that were
# already listed on the legacy 2026 tab (L6:N16), de-duplicated.
HOLIDAYS = [
    (dt.date(2026, 1, 1), "New Year's Day", "Australia (national)"),
    (dt.date(2026, 1, 26), "Australia Day", "Australia (national)"),
    (dt.date(2026, 4, 3), "Good Friday", "Australia (national)"),
    (dt.date(2026, 4, 6), "Easter Monday", "Australia (national)"),
    (dt.date(2026, 4, 25), "ANZAC Day (Saturday)", "Australia (national)"),
    (dt.date(2026, 5, 4), "Labour Day", "QLD"),
    (dt.date(2026, 6, 8), "King's Birthday", "NSW"),
    (dt.date(2026, 10, 5), "King's Birthday", "QLD"),
    (dt.date(2026, 10, 5), "Labour Day", "NSW"),
    (dt.date(2026, 12, 25), "Christmas Day", "Australia (national)"),
    (dt.date(2026, 12, 28), "Boxing Day (observed)", "Australia (national)"),
]

PARAPLANNERS = ["Mayank", "Riya", "Bhakti", "Vijay"]

WEEK_ANCHOR = dt.date(2025, 12, 29)   # week 1 Monday on the legacy 2026 tab

LEAVE_MARKERS = ("away", "on al", "not available", "annual leave")
LEAVE_EXACT = ("al", "a/l", "on leave")
# Placeholders the old tab used for "this slot is empty".
SKIP_EXACT = ("n/a", "na", "-", "?", "tbc")


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------
def put(ws, coord, value, font=BODY, fill=None, fmt=None, align=None,
        border=False, wrap=False):
    c = ws[coord]
    c.value = value
    c.font = font
    if fill is not None:
        c.fill = fill
    if fmt:
        c.number_format = fmt
    if align or wrap:
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        c.border = BOX
    return c


def header_row(ws, row, headers, start_col=1, fill=FILL_HDR, font=HDR):
    for i, text in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=text)
        c.font = font
        c.fill = fill
        c.border = BOX
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[row].height = 30


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def day_suffix_formula(date_ref):
    """'27th' style day-of-month text for a date reference."""
    d = f"DAY({date_ref})"
    return (f'{d}&IF(OR({d}=1,{d}=21,{d}=31),"st",'
            f'IF(OR({d}=2,{d}=22),"nd",IF(OR({d}=3,{d}=23),"rd","th")))')


# --------------------------------------------------------------------------
# Migration: read the legacy workbook
# --------------------------------------------------------------------------
def parse_legacy(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["SOA Allocation 2026"]

    week = None
    week_notes = {}
    allocations = []
    leave = OrderedDict()          # (associate, week) -> reason

    for r in range(2, ws.max_row + 1):
        a, b, c, d, e, f, g, h, i, j = (ws.cell(r, k).value for k in range(1, 11))
        if isinstance(a, (int, float)):
            week = int(a)
            if isinstance(d, str) and d.strip():
                week_notes[week] = d.strip()
            continue
        if week is None:
            continue

        client = str(d).strip() if d is not None else ""
        assoc = str(e).strip() if e is not None else ""
        if not client and not assoc:
            continue
        low = client.lower().strip()
        if any(m in low for m in LEAVE_MARKERS) or low in LEAVE_EXACT:
            if assoc:
                leave.setdefault((assoc, week), client)
            continue
        if low in SKIP_EXACT:
            continue
        if not client or not assoc:
            # An empty holding slot ("associate has a gap") - nothing to carry.
            continue

        allocations.append({
            "week": week,
            "client": client,
            "assoc": assoc,
            "type": str(f).strip() if f else "",
            "division": str(g).strip() if g else "",
            "adviser": str(h).strip() if h else "",
            "soa_date": i,
            "notes": str(j).strip() if j else "",
        })

    history_2025 = read_block(wb["SOA Allocations 2025"], last_col=12)
    peer_history = read_peer_history(wb["Peer Review Allocations"])
    cancelled = read_block(wb["RS_Cancelled SOAs"], last_col=8)
    return allocations, week_notes, leave, history_2025, peer_history, cancelled


def read_block(ws, last_col):
    rows = []
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, last_col + 1)]
        if any(v is not None for v in vals):
            rows.append((r, vals))
    return rows


def read_peer_history(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        who = ws.cell(r, 1).value
        what = ws.cell(r, 2).value
        note = ws.cell(r, 6).value
        if who is None and what is None:
            continue
        rows.append((who, what, note))
    return rows


APPT_MAP = [
    ("tbc", "TBC"), ("tba", "TBC"),
    ("no appt", "No appt"), ("no apt", "No appt"), ("no app", "No appt"),
    ("no - appt", "No appt"),
    ("not booked", "Not booked yet"), ("no date", "Not booked yet"),
    ("asap", "Not booked yet"),
    ("n/a", "N/A"), ("na", "N/A"), ("-", "N/A"),
]


def normalise(alloc):
    """Split the legacy mixed columns into clean, typed fields."""
    out = dict(alloc)
    notes = [alloc["notes"]] if alloc["notes"] else []

    # "Client Type" was doubling as a status column.
    ctype, progress = alloc["type"], ""
    if ctype == "Meeting Cancelled":
        ctype, progress = "", "Cancelled"
    elif ctype in ("Reschedule", "Rescheduled"):
        ctype, progress = "", "Rescheduled"
    elif ctype == "On Hold":
        ctype, progress = "", "On hold"
    out["type"], out["progress"] = ctype, progress

    # "SOA Date" held either a real date or free text.
    raw = alloc["soa_date"]
    date_val, appt = None, ""
    if isinstance(raw, dt.datetime):
        date_val, appt = raw.date(), "Booked"
    elif isinstance(raw, dt.date):
        date_val, appt = raw, "Booked"
    elif isinstance(raw, str) and raw.strip():
        txt = raw.strip()
        low = txt.lower().rstrip(" .")
        matched = ""
        for needle, mapped in APPT_MAP:
            if low == needle or low.startswith(needle):
                appt, matched = mapped, needle
                break
        if not appt:
            appt = "Not booked yet"
        # Anything the old cell said beyond the plain marker is kept in Notes
        # rather than thrown away.
        if low != matched:
            notes.append(f"[SOA date field said: {txt}]")
    out["soa_date"], out["appt"] = date_val, appt
    out["notes"] = " ".join(notes).strip()
    return out


# --------------------------------------------------------------------------
# Sheet builders
# --------------------------------------------------------------------------
def build_setup(wb, leave_rows):
    ws = wb.create_sheet("Setup")
    ws.sheet_properties.tabColor = "FF1F3864"
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 24, "B": 24, "C": 18, "D": 11, "E": 14, "F": 16, "G": 46,
                "H": 3, "I": 13, "J": 13, "K": 22, "L": 22})

    put(ws, "A1", "SETUP", H1)
    put(ws, "A2", "Edit the yellow cells only. Everything else in the workbook "
                  "updates itself from what is here.", MUTED)

    # 1. Plan year -------------------------------------------------------
    put(ws, "A4", "1. Plan year", H2)
    rows = [
        ("Year label", 2026, None),
        ("Week 1 Monday", WEEK_ANCHOR, DATE_FMT),
        ("Weeks to generate", N_WEEK_ROWS, None),
    ]
    for i, (label, value, fmt) in enumerate(rows):
        r = 5 + i
        put(ws, f"A{r}", label, BODY_B)
        put(ws, f"B{r}", value, BODY, FILL_IN, fmt, border=True)
    put(ws, "A8", "Today", BODY_B)
    put(ws, "B8", "=TODAY()", CALC, FILL_CALC, DATE_FMT, border=True)
    put(ws, "A9", "Current week no.", BODY_B)
    put(ws, "B9", '=IFERROR(INDEX(WeekNo,MATCH($B$8,WeekStart,1)),"")',
        CALC, FILL_CALC, border=True)
    put(ws, "C6", "Monday of the first week of the plan year.", MUTED)
    put(ws, "C7", "The Weeks tab builds this many weeks automatically.", MUTED)

    # 2. Roster ----------------------------------------------------------
    put(ws, "A11", "2. Team roster", H2)
    put(ws, "C11", "Add a person here and every dropdown, the dashboard and "
                   "the peer review rotation pick them up.", MUTED)
    header_row(ws, 12, ["Short name (used in Allocations)", "Full name",
                        "Role", "Active?", "SOAs per week",
                        "In peer review rotation?", "Notes"])
    header_row(ws, 12, ["Active rank", "Rotation rank", "Active list",
                        "Rotation order"], start_col=9, fill=FILL_HELP,
               font=HELP_HDR)
    put(ws, "I11", "helper columns - do not edit", MUTED)

    first, last = 13, 13 + 29
    for i in range(30):
        r = first + i
        data = ROSTER[i] if i < len(ROSTER) else ("", "", "", "", None, "", "")
        for col, val in zip("ABCDEFG", data):
            put(ws, f"{col}{r}", val if val != "" else None, BODY, FILL_IN,
                border=True, align="left" if col in "ABCG" else "center")
        put(ws, f"I{r}", f'=IF($D{r}="","",IF($D{r}="Y",COUNTIF($D${first}:$D{r},"Y"),""))',
            HELP_HDR, FILL_HELP, border=True)
        put(ws, f"J{r}", f'=IF(AND($D{r}="Y",$F{r}="Y"),'
                         f'SUMPRODUCT(($D${first}:$D{r}="Y")*($F${first}:$F{r}="Y")),"")',
            HELP_HDR, FILL_HELP, border=True)
        put(ws, f"K{r}", f'=IFERROR(INDEX($A${first}:$A${last},'
                         f'MATCH(ROW()-{first - 1},$I${first}:$I${last},0)),"")',
            HELP_HDR, FILL_HELP, border=True)
        put(ws, f"L{r}", f'=IFERROR(INDEX($A${first}:$A${last},'
                         f'MATCH(ROW()-{first - 1},$J${first}:$J${last},0)),"")',
            HELP_HDR, FILL_HELP, border=True)

    yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(yn)
    yn.add(f"D{first}:D{last}")
    yn.add(f"F{first}:F{last}")

    # 3. Dropdown lists --------------------------------------------------
    lists_row = last + 2
    put(ws, f"A{lists_row}", "3. Dropdown lists", H2)
    put(ws, f"C{lists_row}", "Type into a blank cell in any list and the "
                             "dropdown grows.", MUTED)
    list_specs = [("A", "Advisers", ADVISERS), ("C", "Divisions", DIVISIONS),
                  ("E", "Client types", CLIENT_TYPES), ("G", "Progress", PROGRESS),
                  ("I", "Appt status", APPT_STATUS),
                  ("K", "Peer review status", REVIEW_STATUS)]
    lh = lists_row + 1
    lfirst, llast = lh + 1, lh + 20
    for col, title, values in list_specs:
        put(ws, f"{col}{lh}", title, HDR, FILL_HDR, align="center", border=True)
        for i in range(20):
            put(ws, f"{col}{lfirst + i}",
                values[i] if i < len(values) else None, BODY, FILL_IN,
                border=True)

    # 4. Division caps ---------------------------------------------------
    caps_row = llast + 2
    put(ws, f"A{caps_row}", "4. Division caps per week", H2)
    put(ws, f"D{caps_row}", "Carried over from the Assoc./Imp. Caps block on "
                            "the 2025 tab.", MUTED)
    header_row(ws, caps_row + 1, ["Division", "SOA cap per week",
                                  "Implementation cap per week"])
    cfirst = caps_row + 2
    for i in range(10):
        r = cfirst + i
        row = DIVISION_CAPS[i] if i < len(DIVISION_CAPS) else ("", None, None)
        for col, val in zip("ABC", row):
            put(ws, f"{col}{r}", val if val != "" else None, BODY, FILL_IN,
                border=True, align="center" if col != "A" else "left")
    clast = cfirst + 9

    # 5. Public holidays -------------------------------------------------
    hol_row = clast + 2
    put(ws, f"A{hol_row}", "5. Public holidays", H2)
    put(ws, f"D{hol_row}", "Feeds the Weeks tab: holidays in a week reduce that "
                           "week's working days. Check the state list applies "
                           "to your team.", MUTED)
    header_row(ws, hol_row + 1, ["Date", "Holiday", "State / region"])
    header_row(ws, hol_row + 1, ["Row no."], start_col=5, fill=FILL_HELP,
               font=HELP_HDR)
    hfirst = hol_row + 2
    for i in range(40):
        r = hfirst + i
        row = HOLIDAYS[i] if i < len(HOLIDAYS) else (None, None, None)
        put(ws, f"A{r}", row[0], BODY, FILL_IN, DATE_FMT, border=True)
        put(ws, f"B{r}", row[1], BODY, FILL_IN, border=True)
        put(ws, f"C{r}", row[2], BODY, FILL_IN, border=True)
        # 0 rather than "" - this column is multiplied inside SUMPRODUCT, and
        # text in an arithmetic array gives #VALUE!.
        put(ws, f"E{r}", f'=IF($A{r}="",0,{i + 1})', HELP_HDR, FILL_HELP,
            border=True)
    hlast = hfirst + 39

    # 6. Leave register --------------------------------------------------
    lv_row = hlast + 2
    put(ws, f"A{lv_row}", "6. Leave and unavailable", H2)
    put(ws, f"E{lv_row}", "Anyone listed here is flagged on the Allocations "
                          "tab if they get work in those weeks, and their "
                          "capacity comes out of that week's total.", MUTED)
    header_row(ws, lv_row + 1, ["Associate", "From", "To", "Reason"])
    header_row(ws, lv_row + 1, ["Capacity"], start_col=6, fill=FILL_HELP,
               font=HELP_HDR)
    vfirst = lv_row + 2
    n_leave = max(120, len(leave_rows) + 60)
    for i in range(n_leave):
        r = vfirst + i
        row = leave_rows[i] if i < len(leave_rows) else (None, None, None, None)
        put(ws, f"A{r}", row[0], BODY, FILL_IN, border=True)
        put(ws, f"B{r}", row[1], BODY, FILL_IN, DATE_FMT, border=True)
        put(ws, f"C{r}", row[2], BODY, FILL_IN, DATE_FMT, border=True)
        put(ws, f"D{r}", row[3], BODY, FILL_IN, border=True)
        put(ws, f"F{r}", f'=IF($A{r}="",0,IFERROR(INDEX(RosterCap,'
                         f'MATCH($A{r},RosterName,0)),0))', HELP_HDR,
            FILL_HELP, border=True)
    vlast = vfirst + n_leave - 1

    dv_assoc_leave = DataValidation(type="list", formula1="=Associates",
                                   allow_blank=True)
    ws.add_data_validation(dv_assoc_leave)
    dv_assoc_leave.add(f"A{vfirst}:A{vlast}")

    # 7. Reference -------------------------------------------------------
    ref_row = vlast + 2
    put(ws, f"A{ref_row}", "7. Paraplanning team (reference only)", H2)
    for i, name in enumerate(PARAPLANNERS):
        put(ws, f"A{ref_row + 1 + i}", name, BODY, FILL_IN, border=True)
    put(ws, f"C{ref_row}", "Carried across from the note block on the legacy "
                           "2026 tab (L20:O20).", MUTED)

    # Counters used by the round robin -----------------------------------
    put(ws, "E5", "Active associates", BODY_B)
    put(ws, "F5", f'=COUNTIF($D${first}:$D${last},"Y")', CALC, FILL_CALC,
        border=True)
    put(ws, "E6", "In peer review rotation", BODY_B)
    put(ws, "F6", f'=SUMPRODUCT(($D${first}:$D${last}="Y")*'
                  f'($F${first}:$F${last}="Y"))', CALC, FILL_CALC, border=True)
    put(ws, "E7", "Team SOA capacity per week", BODY_B)
    put(ws, "F7", f'=SUMIF($D${first}:$D${last},"Y",$E${first}:$E${last})',
        CALC, FILL_CALC, border=True)

    ws.freeze_panes = "A5"

    return {
        "roster_first": first, "roster_last": last,
        "list_first": lfirst, "list_last": llast, "list_cols": list_specs,
        "caps_first": cfirst, "caps_last": clast,
        "hol_first": hfirst, "hol_last": hlast,
        "leave_first": vfirst, "leave_last": vlast,
    }


def build_weeks(wb, week_notes):
    ws = wb.create_sheet("Weeks")
    ws.sheet_properties.tabColor = "FF2E75B6"
    widths(ws, {"A": 8, "B": 12, "C": 12, "D": 8, "E": 16, "F": 8, "G": 30,
                "H": 8, "I": 44, "J": 11, "K": 11, "L": 11, "M": 11, "N": 10})
    header_row(ws, 1, ["Week", "Monday", "Friday", "Month", "Week label",
                       "Public hols", "Which holiday", "Working days",
                       "Week note (type here)", "SOAs allocated",
                       "Team capacity", "Capacity on leave",
                       "Capacity available", "Spare"])
    for r in range(2, N_WEEK_ROWS + 2):
        i = r - 1
        put(ws, f"A{r}", f'=IF({i}>Setup!$B$7,"",{i})',
            CALC, FILL_CALC, border=True, align="center")
        put(ws, f"B{r}", f'=IF($A{r}="","",Setup!$B$6+7*($A{r}-1))', CALC,
            FILL_CALC, DATE_FMT, border=True)
        put(ws, f"C{r}", f'=IF($A{r}="","",$B{r}+4)', CALC, FILL_CALC,
            DATE_FMT, border=True)
        put(ws, f"D{r}", f'=IF($A{r}="","",UPPER(TEXT($B{r},"mmm")))', CALC,
            FILL_CALC, border=True, align="center")
        put(ws, f"E{r}", f'=IF($A{r}="","",{day_suffix_formula(f"$B{r}")}&" - "'
                         f'&{day_suffix_formula(f"$C{r}")})', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"F{r}", f'=IF($A{r}="","",SUMPRODUCT((HolDate>=$B{r})*'
                         f'(HolDate<=$C{r})))', CALC, FILL_CALC, border=True,
            align="center")
        put(ws, f"G{r}", f'=IF($F{r}=0,"",IF($F{r}>1,$F{r}&" public holidays",'
                         f'IFERROR(INDEX(HolName,SUMPRODUCT((HolDate>=$B{r})*'
                         f'(HolDate<=$C{r})*HolIdx)),"")))', CALC, FILL_CALC,
            border=True)
        put(ws, f"H{r}", f'=IF($A{r}="","",5-$F{r})', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"I{r}", week_notes.get(i), BODY, FILL_IN, border=True,
            wrap=True)
        put(ws, f"J{r}", f'=IF($A{r}="","",SUMPRODUCT((AllocWeek=$A{r})*'
                         f'(AllocClient<>"")))', CALC, FILL_CALC, border=True,
            align="center")
        put(ws, f"K{r}", f'=IF($A{r}="","",Setup!$F$7)', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"L{r}", f'=IF($A{r}="","",SUMPRODUCT((LeaveFrom<=$C{r})*'
                         f'(LeaveTo>=$B{r})*LeaveCap))', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"M{r}", f'=IF($A{r}="","",$K{r}-$L{r})', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"N{r}", f'=IF($A{r}="","",$M{r}-$J{r})', CALC, FILL_CALC,
            border=True, align="center")

    # Per-division counts per week. The dashboard reads this block, which
    # keeps every formula a plain COUNTIFS instead of an array.
    for i in range(7):
        col = get_column_letter(17 + i)
        widths(ws, {col: 10})
        put(ws, f"{col}1", f'=IFERROR(INDEX(CapDivision,{i + 1}),"")',
            HELP_HDR, FILL_HELP, border=True, align="center")
        for r in range(2, N_WEEK_ROWS + 2):
            put(ws, f"{col}{r}", f'=IF(OR($A{r}="",{col}$1=""),"",'
                                 f'COUNTIFS(AllocWeek,$A{r},AllocDivision,'
                                 f'{col}$1))', HELP_HDR, FILL_HELP,
                border=True, align="center")
        ws.column_dimensions[col].hidden = True

    rng = f"A2:N{N_WEEK_ROWS + 1}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'AND($B2<>"",TODAY()>=$B2,TODAY()<=$C2)'], fill=FILL_NOW))
    ws.conditional_formatting.add(f"N2:N{N_WEEK_ROWS + 1}", FormulaRule(
        formula=['AND($N2<>"",$N2<0)'], fill=FILL_BAD))
    ws.conditional_formatting.add(f"N2:N{N_WEEK_ROWS + 1}", FormulaRule(
        formula=['AND($N2<>"",$N2=0)'], fill=FILL_WARN))
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:N{N_WEEK_ROWS + 1}"


ALLOC_HEADERS = [
    ("A", "Week", 8), ("B", "Month", 8), ("C", "Week label", 15),
    ("D", "Week Monday", 13), ("E", "Client", 38), ("F", "Associate", 14),
    ("G", "# for assoc. that week", 9), ("H", "Client type", 14),
    ("I", "Division", 11), ("J", "Adviser", 14), ("K", "SOA / appt date", 14),
    ("L", "Appt status", 13), ("M", "Progress", 17),
    ("N", "Peer reviewer (auto)", 15), ("O", "Reviewer override", 15),
    ("P", "Peer reviewer", 15), ("Q", "Checks", 34), ("R", "Notes", 50),
    ("S", "SOA folder link", 30), ("T", "Peer review status", 15),
    ("U", "Reviewed on", 13), ("V", "Reviewer comments", 40),
]
INPUT_COLS = "AEFHIJKLMORSTUV"
CALC_COLS = "BCDGNPQ"
HELPER_COLS = [("X", "Seq no."), ("Y", "Checks (raw)"),
               ("Z", "Week|assoc|# key"), ("AA", "Reviewer rank"),
               ("AB", "Reviewer|rank key"), ("AC", "Author's nth SOA")]


def build_allocations(wb, allocations):
    ws = wb.create_sheet("Allocations")
    ws.sheet_properties.tabColor = "FF548235"
    widths(ws, {col: w for col, _, w in ALLOC_HEADERS})
    widths(ws, {"W": 3, "X": 9, "Y": 40, "Z": 20, "AA": 9, "AB": 20,
                "AC": 9})

    header_row(ws, 1, [h for _, h, _ in ALLOC_HEADERS])
    header_row(ws, 1, [h for _, h in HELPER_COLS], start_col=24,
               fill=FILL_HELP, font=HELP_HDR)

    last = LAST_ALLOC_ROW
    for r in range(2, last + 1):
        idx = r - 2
        a = allocations[idx] if idx < len(allocations) else None

        put(ws, f"A{r}", a["week"] if a else None, BODY, FILL_IN,
            border=True, align="center")
        put(ws, f"B{r}", f'=IF($A{r}="","",IFERROR(INDEX(WeekMonth,'
                         f'MATCH($A{r},WeekNo,0)),"?"))', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"C{r}", f'=IF($A{r}="","",IFERROR(INDEX(WeekLabel,'
                         f'MATCH($A{r},WeekNo,0)),"?"))', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"D{r}", f'=IF($A{r}="","",IFERROR(INDEX(WeekStart,'
                         f'MATCH($A{r},WeekNo,0)),""))', CALC, FILL_CALC,
            DATE_FMT, border=True)
        put(ws, f"E{r}", a["client"] if a else None, BODY, FILL_IN, border=True)
        put(ws, f"F{r}", a["assoc"] if a else None, BODY, FILL_IN, border=True)
        put(ws, f"G{r}", f'=IF(OR($E{r}="",$F{r}=""),"",'
                         f'SUMPRODUCT(($A$2:$A{r}=$A{r})*($F$2:$F{r}=$F{r})*'
                         f'($E$2:$E{r}<>"")))',
            CALC, FILL_CALC, border=True, align="center")
        put(ws, f"H{r}", a["type"] or None if a else None, BODY, FILL_IN,
            border=True)
        put(ws, f"I{r}", a["division"] or None if a else None, BODY, FILL_IN,
            border=True)
        put(ws, f"J{r}", a["adviser"] or None if a else None, BODY, FILL_IN,
            border=True)
        put(ws, f"K{r}", a["soa_date"] if a else None, BODY, FILL_IN,
            DATE_FMT, border=True)
        put(ws, f"L{r}", a["appt"] or None if a else None, BODY, FILL_IN,
            border=True)
        put(ws, f"M{r}", a["progress"] or None if a else None, BODY, FILL_IN,
            border=True)
        # Round robin. Nobody reviews their own SOA, so each associate's own
        # SOAs are dealt out in turn across everyone else in the rotation
        # (their own count drives the turn, offset by their position so two
        # authors do not start on the same reviewer). Anyone not in the
        # rotation falls back to the whole list in file order.
        author = f'MATCH($F{r},Reviewers,0)'
        slot = f'MOD($AC{r}-1+{author},MAX(ReviewerCount-1,1))+1'
        put(ws, f"N{r}", f'=IF(OR($E{r}="",$F{r}="",ReviewerCount=0),"",'
                         f'IFERROR(INDEX(Reviewers,IF(ISNA({author}),'
                         f'MOD($X{r}-1,ReviewerCount)+1,'
                         f'IF({slot}<{author},{slot},{slot}+1))),""))',
            CALC, FILL_CALC, border=True)
        put(ws, f"O{r}", None, BODY, FILL_IN, border=True)
        put(ws, f"P{r}", f'=IF($O{r}<>"",$O{r},$N{r})', CALC, FILL_CALC,
            border=True)
        put(ws, f"Q{r}", f'=IF($E{r}="","",IF($Y{r}="","OK",MID($Y{r},3,300)))',
            CALC, FILL_CALC, border=True, wrap=True)
        put(ws, f"R{r}", a["notes"] or None if a else None, BODY, FILL_IN,
            border=True, wrap=True)
        for col in "STUV":
            put(ws, f"{col}{r}", None, BODY, FILL_IN, border=True,
                fmt=DATE_FMT if col == "U" else None)

        # helpers
        put(ws, f"X{r}", f'=IF(OR($E{r}="",$A{r}=""),"",'
                         f'SUMPRODUCT((AllocWeek<>"")*(AllocWeek<$A{r})*'
                         f'(AllocClient<>""))+SUMPRODUCT(($A$2:$A{r}=$A{r})*'
                         f'($E$2:$E{r}<>"")))',
            HELP_HDR, FILL_HELP, border=True)
        cap = f'IFERROR(INDEX(RosterCap,MATCH($F{r},RosterName,0)),99)'
        put(ws, f"Y{r}", (
            f'=IF($E{r}="","",'
            f'IF($F{r}="","; no associate","")'
            f'&IF(AND($F{r}<>"",COUNTIFS(RosterName,$F{r})=0),'
            f'"; associate not in roster","")'
            f'&IF(AND($F{r}<>"",$D{r}<>"",COUNTIFS(LeaveWho,$F{r},'
            f'LeaveFrom,"<="&($D{r}+4),LeaveTo,">="&$D{r})>0),'
            f'"; on leave that week","")'
            f'&IF(AND($F{r}<>"",SUMPRODUCT((AllocWeek=$A{r})*'
            f'(AllocAssoc=$F{r})*(AllocClient<>""))>{cap}),'
            f'"; over capacity","")'
            f'&IF(AND(ISNUMBER($K{r}),$D{r}<>"",$K{r}<$D{r}),'
            f'"; SOA date before the week","")'
            f'&IF(AND(ISNUMBER($K{r}),$D{r}<>"",$K{r}>$D{r}+90),'
            f'"; SOA date over 90 days out","")'
            f'&IF(AND($P{r}<>"",$P{r}=$F{r}),"; reviewer is the author","")'
            f'&IF(SUMPRODUCT((AllocWeek=$A{r})*(AllocClient=$E{r}))>1,'
            f'"; duplicate in this week","")'
            f'&IF($J{r}="","; no adviser","")'
            f'&IF($A{r}="","; no week","")'
            f')'), HELP_HDR, FILL_HELP, border=True)
        put(ws, f"Z{r}", f'=IF($E{r}="","",$A{r}&"|"&$F{r}&"|"&$G{r})',
            HELP_HDR, FILL_HELP, border=True)
        put(ws, f"AA{r}", f'=IF(OR($E{r}="",$P{r}=""),"",'
                          f'SUMPRODUCT(($P$2:$P{r}=$P{r})*'
                          f'($E$2:$E{r}<>"")))',
            HELP_HDR, FILL_HELP, border=True)
        put(ws, f"AB{r}", f'=IF($AA{r}="","",$P{r}&"|"&$AA{r})',
            HELP_HDR, FILL_HELP, border=True)
        put(ws, f"AC{r}", f'=IF(OR($E{r}="",$F{r}=""),"",'
                          f'SUMPRODUCT(($F$2:$F{r}=$F{r})*($E$2:$E{r}<>"")))',
            HELP_HDR, FILL_HELP, border=True)

    # Dropdowns
    dvs = [
        ("=Associates", "F"), ("=Associates", "O"), ("=Advisers", "J"),
        ("=Divisions", "I"), ("=ClientTypes", "H"), ("=ProgressList", "M"),
        ("=ApptList", "L"), ("=ReviewStatusList", "T"), ("=WeekNo", "A"),
    ]
    for formula, col in dvs:
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{last}")

    ws.conditional_formatting.add(f"A2:V{last}", FormulaRule(
        formula=['AND($D2<>"",TODAY()>=$D2,TODAY()<=$D2+4)'], fill=FILL_NOW))
    ws.conditional_formatting.add(f"Q2:Q{last}", FormulaRule(
        formula=['AND($Q2<>"",$Q2<>"OK")'], fill=FILL_WARN))
    ws.conditional_formatting.add(f"M2:M{last}", FormulaRule(
        formula=['OR($M2="Cancelled",$M2="On hold",$M2="Rescheduled")'],
        fill=FILL_BAD))
    ws.conditional_formatting.add(f"M2:M{last}", FormulaRule(
        formula=['$M2="Completed"'], fill=FILL_OK))

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:V{last}"
    for col in ("X", "Y", "Z", "AA", "AB", "AC"):
        ws.column_dimensions[col].hidden = True


def build_week_view(wb):
    ws = wb.create_sheet("Week View")
    ws.sheet_properties.tabColor = "FFED7D31"
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 22})
    for i in range(N_ASSOC_COLS):
        widths(ws, {get_column_letter(2 + i): 26})

    put(ws, "A1", "WEEK VIEW", H1)
    put(ws, "A2", "Type a week number in B3. Everything below is built from "
                  "the Allocations tab - nothing here needs maintaining.", MUTED)
    put(ws, "A3", "Week number", BODY_B)
    put(ws, "B3", 31, BODY_B, FILL_IN, border=True, align="center")
    put(ws, "C3", '=IF($B$3="","",IFERROR(INDEX(WeekMonth,MATCH($B$3,WeekNo,0))'
                  '&"  "&INDEX(WeekLabel,MATCH($B$3,WeekNo,0)),"week not found"))',
        H2)
    put(ws, "A4", "This week is week", MUTED)
    put(ws, "B4", '=IFERROR(INDEX(WeekNo,MATCH(TODAY(),WeekStart,1)),"")',
        CALC, FILL_CALC, border=True, align="center")
    put(ws, "C4", '=IFERROR(IF(INDEX(WeekNote,MATCH($B$3,WeekNo,0))="","",'
                  '"Week note: "&INDEX(WeekNote,MATCH($B$3,WeekNo,0))),"")',
        MUTED)

    put(ws, "A6", "Associate", HDR, FILL_HDR, border=True, align="center")
    put(ws, "A7", "Allocated / capacity", BODY_B, FILL_SUB, border=True,
        align="center")
    put(ws, "A8", "On leave this week?", BODY_B, FILL_SUB, border=True,
        align="center")
    for i in range(N_ASSOC_COLS):
        col = get_column_letter(2 + i)
        put(ws, f"{col}6", f'=IFERROR(INDEX(Associates,{i + 1}),"")', HDR,
            FILL_HDR, border=True, align="center")
        put(ws, f"{col}7", f'=IF({col}$6="","",SUMPRODUCT((AllocWeek=$B$3)*'
                           f'(AllocAssoc={col}$6)*(AllocClient<>""))&" / "'
                           f'&IFERROR(INDEX(RosterCap,'
                           f'MATCH({col}$6,RosterName,0)),"?"))', CALC,
            FILL_SUB, border=True, align="center")
        put(ws, f"{col}8", f'=IF({col}$6="","",IF(SUMPRODUCT((LeaveWho={col}$6)*'
                           f'(LeaveFrom<=IFERROR(INDEX(WeekEnd,MATCH($B$3,WeekNo,0)),0))*'
                           f'(LeaveTo>=IFERROR(INDEX(WeekStart,MATCH($B$3,WeekNo,0)),0)))>0,'
                           f'"ON LEAVE",""))', CALC, FILL_SUB, border=True,
            align="center")
        for k in range(1, 11):
            r = 8 + k
            put(ws, f"{col}{r}", f'=IFERROR(INDEX(AllocClient,MATCH($B$3&"|"&'
                                 f'{col}$6&"|"&{k},AllocKey,0)),"")', BODY,
                FILL_CALC, border=True, wrap=True)
    for k in range(1, 11):
        put(ws, f"A{8 + k}", f"SOA {k}", BODY_B, FILL_SUB, border=True,
            align="center")

    rng = f"B9:{get_column_letter(1 + N_ASSOC_COLS)}18"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=['B9<>""'], fill=FILL_OK))
    ws.conditional_formatting.add(
        f"B8:{get_column_letter(1 + N_ASSOC_COLS)}8",
        FormulaRule(formula=['B8="ON LEAVE"'], fill=FILL_BAD))

    put(ws, "A21", "Week summary", H2)
    summary = [
        ("SOAs allocated",
         '=IF($B$3="","",SUMPRODUCT((AllocWeek=$B$3)*(AllocClient<>"")))'),
        ("Capacity available", '=IFERROR(INDEX(WeekAvail,MATCH($B$3,WeekNo,0)),"")'),
        ("Spare capacity", '=IFERROR(INDEX(WeekSpare,MATCH($B$3,WeekNo,0)),"")'),
        ("New business", '=IF($B$3="","",COUNTIFS(AllocWeek,$B$3,AllocType,"New Business"))'),
        ("Existing clients", '=IF($B$3="","",COUNTIFS(AllocWeek,$B$3,AllocType,"Existing"))'),
        ("Rows with a check to look at",
         '=IF($B$3="","",SUMPRODUCT((AllocWeek=$B$3)*(AllocClient<>"")*'
         '(AllocChecks<>"OK")))'),
        ("Public holidays this week",
         '=IFERROR(INDEX(WeekHols,MATCH($B$3,WeekNo,0)),"")'),
    ]
    for i, (label, formula) in enumerate(summary):
        r = 22 + i
        put(ws, f"A{r}", label, BODY_B, FILL_SUB, border=True)
        put(ws, f"B{r}", formula, CALC, FILL_CALC, border=True, align="center")

    put(ws, "D21", "SOAs per division this week", H2)
    for i, div in enumerate(DIVISIONS):
        r = 22 + i
        put(ws, f"D{r}", f'=IFERROR(INDEX(CapDivision,{i + 1}),"")', BODY_B,
            FILL_SUB, border=True)
        put(ws, f"E{r}", f'=IF(OR($B$3="",$D{r}=""),"",SUMPRODUCT((AllocWeek=$B$3)*'
                         f'(AllocDivision=$D{r})*(AllocClient<>""))&" / "'
                         f'&IFERROR(INDEX(CapSOA,'
                         f'MATCH($D{r},CapDivision,0)),"?"))', CALC, FILL_CALC,
            border=True, align="center")
    ws.freeze_panes = "B6"


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = "FF7030A0"
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 8, "B": 12, "C": 15})
    for i in range(N_ASSOC_COLS):
        widths(ws, {get_column_letter(4 + i): 12})
    tot_col = get_column_letter(4 + N_ASSOC_COLS)
    cap_col = get_column_letter(5 + N_ASSOC_COLS)
    spare_col = get_column_letter(6 + N_ASSOC_COLS)
    widths(ws, {tot_col: 10, cap_col: 12, spare_col: 10})

    put(ws, "A1", "LOAD DASHBOARD", H1)
    put(ws, "A2", "SOAs allocated per associate per week. Red = over that "
                  "person's weekly number, amber = at it, green = has room, "
                  "blue row = the current week.", MUTED)

    header_row(ws, 4, ["Week", "Monday", "Week label"])
    for i in range(N_ASSOC_COLS):
        col = get_column_letter(4 + i)
        put(ws, f"{col}4", f'=IFERROR(INDEX(Associates,{i + 1}),"")', HDR,
            FILL_HDR, border=True, align="center")
    header_row(ws, 4, ["Total", "Capacity", "Spare"], start_col=4 + N_ASSOC_COLS)

    put(ws, "C5", "Capacity per week", BODY_B, FILL_SUB, border=True,
        align="right")
    for i in range(N_ASSOC_COLS):
        col = get_column_letter(4 + i)
        put(ws, f"{col}5", f'=IF({col}$4="","",IFERROR(INDEX(RosterCap,'
                           f'MATCH({col}$4,RosterName,0)),""))', BODY_B,
            FILL_SUB, border=True, align="center")

    first_row = 6
    for k in range(N_WEEK_ROWS):
        r = first_row + k
        put(ws, f"A{r}", f'=IFERROR(INDEX(WeekNo,{k + 1}),"")', CALC,
            FILL_CALC, border=True, align="center")
        put(ws, f"B{r}", f'=IF($A{r}="","",INDEX(WeekStart,{k + 1}))', CALC,
            FILL_CALC, DATE_FMT, border=True)
        put(ws, f"C{r}", f'=IF($A{r}="","",INDEX(WeekMonth,{k + 1})&" "&'
                         f'INDEX(WeekLabel,{k + 1}))', CALC, FILL_CALC,
            border=True)
        for i in range(N_ASSOC_COLS):
            col = get_column_letter(4 + i)
            put(ws, f"{col}{r}", f'=IF(OR($A{r}="",{col}$4=""),"",'
                                 f'COUNTIFS(AllocWeek,$A{r},AllocAssoc,{col}$4))',
                CALC, FILL_CALC, '0;-0;"-"', border=True, align="center")
        put(ws, f"{tot_col}{r}", f'=IF($A{r}="","",SUMPRODUCT((AllocWeek=$A{r})*'
                                 f'(AllocClient<>"")))', BODY_B, FILL_SUB,
            border=True, align="center")
        put(ws, f"{cap_col}{r}", f'=IF($A{r}="","",INDEX(WeekAvail,{k + 1}))',
            BODY_B, FILL_SUB, border=True, align="center")
        put(ws, f"{spare_col}{r}", f'=IF($A{r}="","",INDEX(WeekSpare,{k + 1}))',
            BODY_B, FILL_SUB, border=True, align="center")
    last_row = first_row + N_WEEK_ROWS - 1

    matrix = f"D{first_row}:{get_column_letter(3 + N_ASSOC_COLS)}{last_row}"
    ws.conditional_formatting.add(matrix, FormulaRule(
        formula=[f'AND(ISNUMBER(D{first_row}),D$5<>"",D{first_row}>D$5)'],
        fill=FILL_BAD))
    ws.conditional_formatting.add(matrix, FormulaRule(
        formula=[f'AND(ISNUMBER(D{first_row}),D$5<>"",D{first_row}=D$5)'],
        fill=FILL_WARN))
    ws.conditional_formatting.add(matrix, FormulaRule(
        formula=[f'AND(ISNUMBER(D{first_row}),D{first_row}>0)'], fill=FILL_OK))
    ws.conditional_formatting.add(
        f"A{first_row}:{spare_col}{last_row}",
        FormulaRule(formula=[f'AND($B{first_row}<>"",TODAY()>=$B{first_row},'
                             f'TODAY()<=$B{first_row}+4)'], fill=FILL_NOW))
    ws.conditional_formatting.add(
        f"{spare_col}{first_row}:{spare_col}{last_row}",
        FormulaRule(formula=[f'AND(ISNUMBER({spare_col}{first_row}),'
                             f'{spare_col}{first_row}<0)'], fill=FILL_BAD))

    # Year to date blocks -------------------------------------------------
    r0 = last_row + 3
    put(ws, f"A{r0}", "Year to date", H1)

    header_row(ws, r0 + 1, ["Associate", "SOAs written", "New business",
                            "Existing", "Completed", "Cancelled / on hold",
                            "Peer reviews done"])
    for i in range(N_ASSOC_COLS):
        r = r0 + 2 + i
        put(ws, f"A{r}", f'=IFERROR(INDEX(Associates,{i + 1}),"")', BODY_B,
            FILL_CALC, border=True)
        cells = [
            f'=IF($A{r}="","",SUMPRODUCT((AllocAssoc=$A{r})*'
            f'(AllocClient<>"")))',
            f'=IF($A{r}="","",COUNTIFS(AllocAssoc,$A{r},AllocType,"New Business"))',
            f'=IF($A{r}="","",COUNTIFS(AllocAssoc,$A{r},AllocType,"Existing"))',
            f'=IF($A{r}="","",COUNTIFS(AllocAssoc,$A{r},AllocProgress,"Completed"))',
            f'=IF($A{r}="","",COUNTIFS(AllocAssoc,$A{r},AllocProgress,"Cancelled")'
            f'+COUNTIFS(AllocAssoc,$A{r},AllocProgress,"On hold"))',
            f'=IF($A{r}="","",SUMPRODUCT((AllocReviewer=$A{r})*'
            f'(AllocClient<>"")))',
        ]
        for j, formula in enumerate(cells):
            put(ws, f"{get_column_letter(2 + j)}{r}", formula, CALC, FILL_CALC,
                border=True, align="center")

    r1 = r0 + 2 + N_ASSOC_COLS + 1
    header_row(ws, r1, ["Division", "SOAs YTD", "Weekly SOA cap",
                        "Busiest week", "Weeks over cap"])
    for i in range(len(DIVISIONS) + 2):
        r = r1 + 1 + i
        put(ws, f"A{r}", f'=IFERROR(INDEX(CapDivision,{i + 1}),"")', BODY_B,
            FILL_CALC, border=True)
        put(ws, f"B{r}", f'=IF($A{r}="","",SUMPRODUCT((AllocDivision=$A{r})*'
                         f'(AllocClient<>"")))', CALC, FILL_CALC, border=True,
            align="center")
        put(ws, f"C{r}", f'=IF($A{r}="","",IFERROR(INDEX(CapSOA,'
                         f'MATCH($A{r},CapDivision,0)),""))', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"D{r}", f'=IF($A{r}="","",IFERROR(MAX(INDEX(WeekDivBlock,0,'
                         f'MATCH($A{r},WeekDivHeads,0))),""))', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"E{r}", f'=IF(OR($A{r}="",$C{r}=""),"",IFERROR(COUNTIF('
                         f'INDEX(WeekDivBlock,0,MATCH($A{r},WeekDivHeads,0)),'
                         f'">"&$C{r}),""))', CALC, FILL_CALC, border=True,
            align="center")

    r2 = r1 + len(DIVISIONS) + 4
    header_row(ws, r2, ["Adviser", "SOAs YTD"])
    for i in range(20):
        r = r2 + 1 + i
        put(ws, f"A{r}", f'=IFERROR(INDEX(Advisers,{i + 1}),"")', BODY_B,
            FILL_CALC, border=True)
        put(ws, f"B{r}", f'=IF($A{r}="","",SUMPRODUCT((AllocAdviser=$A{r})*'
                         f'(AllocClient<>"")))', CALC, FILL_CALC, border=True,
            align="center")
    ws.freeze_panes = "D6"


def build_peer_review(wb):
    ws = wb.create_sheet("Peer Review")
    ws.sheet_properties.tabColor = "FFC00000"
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 6, "B": 8, "C": 34, "D": 14, "E": 14, "F": 13, "G": 34,
                "H": 17, "I": 13, "J": 40})

    put(ws, "A1", "PEER REVIEW - ROUND ROBIN", H1)
    put(ws, "A2", "Worked out automatically on the Allocations tab: each "
                  "associate's SOAs are dealt out in turn across everyone else "
                  "in the order below, so nobody reviews their own work. The "
                  "Assigned column below is the fairness check - if it drifts, "
                  "use the Reviewer override column.", MUTED)
    put(ws, "A3", "SOAs must be uploaded for review at least 24 hours before "
                  "the SOA date to guarantee a check.", BODY_B)

    put(ws, "A5", "Rotation order", H2)
    put(ws, "D5", "Set who is in the rotation in Setup, column F.", MUTED)
    header_row(ws, 6, ["#", "Reviewer", "Assigned", "Outstanding", "Approved",
                       "Reviewed this month"], start_col=1)
    for i in range(N_ASSOC_COLS):
        r = 7 + i
        put(ws, f"A{r}", f'=IF($B{r}="","",{i + 1})', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"B{r}", f'=IFERROR(INDEX(Reviewers,{i + 1}),"")', BODY_B,
            FILL_CALC, border=True)
        put(ws, f"C{r}", f'=IF($B{r}="","",SUMPRODUCT((AllocReviewer=$B{r})*'
                         f'(AllocClient<>"")))', CALC, FILL_CALC, border=True,
            align="center")
        put(ws, f"D{r}", f'=IF($B{r}="","",SUMPRODUCT((AllocReviewer=$B{r})*'
                         f'(AllocClient<>""))-COUNTIFS(AllocReviewer,$B{r},'
                         f'AllocRevStatus,"Approved"))', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"E{r}", f'=IF($B{r}="","",COUNTIFS(AllocReviewer,$B{r},'
                         f'AllocRevStatus,"Approved"))', CALC, FILL_CALC,
            border=True, align="center")
        put(ws, f"F{r}", f'=IF($B{r}="","",COUNTIFS(AllocReviewer,$B{r},'
                         f'AllocReviewedOn,">="&DATE(YEAR(TODAY()),'
                         f'MONTH(TODAY()),1)))', CALC, FILL_CALC,
            border=True, align="center")
    ws.conditional_formatting.add("D7:D18", FormulaRule(
        formula=['AND(ISNUMBER(D7),D7>0)'], fill=FILL_WARN))

    put(ws, "A21", "My review queue", H1)
    put(ws, "A22", "Reviewer", BODY_B)
    put(ws, "B22", '=IFERROR(INDEX(Reviewers,1),"")', BODY_B, FILL_IN,
        border=True)
    put(ws, "D22", "Pick your name, then use the filter on row 24 to hide "
                   "what is already approved. Record the outcome against the "
                   "SOA on the Allocations tab (columns S to V).", MUTED)

    header_row(ws, 24, ["#", "Week", "Client", "Written by", "SOA date",
                        "Appt status", "SOA folder link", "Review status",
                        "Reviewed on", "Comments"])
    key = '$B$22&"|"&$A{r}'
    lookups = [("B", "AllocWeek", None), ("C", "AllocClient", None),
               ("D", "AllocAssoc", None), ("E", "AllocDate", DATE_FMT),
               ("F", "AllocAppt", None), ("G", "AllocLink", None),
               ("H", "AllocRevStatus", None), ("I", "AllocReviewedOn", DATE_FMT),
               ("J", "AllocComments", None)]
    for i in range(120):
        r = 25 + i
        put(ws, f"A{r}", f'=IF(SUMPRODUCT((AllocReviewer=$B$22)*'
                         f'(AllocClient<>""))<{i + 1},"",{i + 1})', CALC, FILL_CALC, border=True,
            align="center")
        for col, name, fmt in lookups:
            put(ws, f"{col}{r}", f'=IF($A{r}="","",IFERROR(INDEX({name},'
                                 f'MATCH({key.format(r=r)},AllocReviewKey,0)),""))',
                CALC, FILL_CALC, fmt, border=True,
                wrap=(col in ("C", "J")))
    ws.conditional_formatting.add("A25:J144", FormulaRule(
        formula=['AND($A25<>"",$H25<>"Approved")'], fill=FILL_WARN))

    dv = DataValidation(type="list", formula1="=Reviewers", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("B22")
    ws.auto_filter.ref = "A24:J144"
    ws.freeze_panes = "A25"


def build_start_here(wb):
    ws = wb.create_sheet("Start Here", 0)
    ws.sheet_properties.tabColor = "FF000000"
    ws.sheet_view.showGridLines = False
    widths(ws, {"A": 3, "B": 30, "C": 96})

    put(ws, "B2", "ASSOCIATE ALLOCATIONS", H1)
    put(ws, "B3", "One register, one calendar, one automatic peer review "
                  "rotation. Type in the yellow cells; grey cells look after "
                  "themselves.", MUTED)

    blocks = [
        ("Weekly routine", [
            ("1. Allocations tab", "Add one row per client file: week number, "
             "client, associate, type, division, adviser, SOA date. Everything "
             "else on the row fills itself in."),
            ("2. Check the Checks column", "It tells you if someone is on "
             "leave that week, over their weekly number, missing an adviser, "
             "double-booked on the same client, or has an SOA date that looks "
             "wrong. \"OK\" means nothing to look at."),
            ("3. Dashboard tab", "Look down the current week's row before you "
             "allocate: red means that person is over their number, green "
             "means they have room, and the Spare column is the whole team's "
             "headroom for the week."),
            ("4. Week View tab", "Type a week number for the old "
             "one-column-per-associate view of that week, plus who is on "
             "leave and how the week sits against the division caps."),
            ("5. Peer Review tab", "Reviewers are already assigned. Pick your "
             "name to see your queue, then record the outcome on the "
             "Allocations tab (columns S to V)."),
        ]),
        ("What is now automatic", [
            ("Week calendar", "The Weeks tab builds every week of the year "
             "from one date (Setup B6) - week numbers, months, \"27th - 31st\" "
             "labels, public holidays and working days. No more copying week "
             "blocks down the sheet."),
            ("Peer review rotation", "Nobody reviews their own SOA: each "
             "associate's files are dealt out in turn across everyone else in "
             "the rotation. Change who is in the rotation with one Y/N on "
             "Setup instead of retyping hundreds of rows, and use the Reviewer "
             "override column to swap a single review."),
            ("Load and capacity", "Counts per associate, per division and per "
             "week come from the register, and leave comes out of that week's "
             "capacity automatically."),
            ("Dropdowns", "Associates, advisers, divisions, client types and "
             "statuses all come from the Setup lists, so adding a person or an "
             "adviser is a one-line change."),
        ]),
        ("Colour key", [
            ("Yellow", "You type here."),
            ("Grey / blue text", "Formula - leave it alone, it will rebuild "
             "itself."),
            ("Light blue row", "The current week."),
            ("Green / amber / red", "Has room / at capacity / over capacity, "
             "or a check to look at."),
        ]),
        ("Tabs", [
            ("Setup", "Roster, dropdown lists, division caps, public holidays, "
             "leave. The only tab that needs maintaining."),
            ("Weeks", "The generated week calendar plus a free-text note per "
             "week."),
            ("Allocations", "The register - one row per client file. Filter and "
             "sort it however you like."),
            ("Week View", "One week at a time, laid out per associate."),
            ("Dashboard", "Load per associate per week, plus year-to-date "
             "totals by associate, division and adviser."),
            ("Peer Review", "Rotation order, fairness counts and your review "
             "queue."),
            ("History 2025 / 2026 / Peer Review History / RS-Cancelled",
             "Read-only copies of the old tabs so nothing is lost."),
        ]),
        ("Things worth knowing", [
            ("Locking a reviewer in", "The auto rotation is based on the order "
             "SOAs appear. Once a week is settled, copy column P for that "
             "week and paste it as values into column O (Reviewer override) if "
             "you want those names frozen."),
            ("Adding rows", "The register has formulas down to row "
             f"{LAST_ALLOC_ROW}. If you need more, select the last row and "
             "drag it down - the formulas copy."),
            ("Holidays", "The Setup list has the national dates plus the "
             "NSW/QLD ones that were already on the old tab. Add anything your "
             "state needs."),
            ("Capacity numbers", "Everyone starts at 3 SOAs a week (2 for "
             "Ikesha), taken from the numbers in the old tab headers. Change "
             "them in Setup column E."),
        ]),
    ]

    r = 5
    for title, items in blocks:
        put(ws, f"B{r}", title, H2, FILL_SUB)
        put(ws, f"C{r}", None, H2, FILL_SUB)
        r += 1
        for label, text in items:
            put(ws, f"B{r}", label, BODY_B, border=True, wrap=True)
            put(ws, f"C{r}", text, BODY, border=True, wrap=True)
            ws.row_dimensions[r].height = 30 if len(text) < 120 else 44
            r += 1
        r += 1

    put(ws, f"B{r}", "Built from Associate_Allocations_14.xlsx. The old tabs "
                     "are kept as read-only history tabs; the original file is "
                     "unchanged.", MUTED)


def build_history(wb, history_2025, peer_history, cancelled):
    # 2025 allocations ---------------------------------------------------
    ws = wb.create_sheet("History 2025")
    ws.sheet_properties.tabColor = "FFA6A6A6"
    widths(ws, {"A": 22, "B": 10, "C": 14, "D": 8, "E": 38, "F": 14, "G": 14,
                "H": 11, "I": 14, "J": 14, "K": 22, "L": 46})
    header_row(ws, 1, ["Week note", "Month", "Week", "#", "Client",
                       "Associate", "Client type", "Division", "Adviser",
                       "SOA date", "Data status", "Notes"])
    out = 2
    for _, vals in history_2025[1:]:
        for i, v in enumerate(vals):
            cell = ws.cell(row=out, column=i + 1, value=v)
            cell.font = BODY
            if isinstance(v, dt.datetime):
                cell.number_format = DATE_FMT
        out += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{out - 1}"

    # Peer review history ------------------------------------------------
    ws = wb.create_sheet("Peer Review History")
    ws.sheet_properties.tabColor = "FFA6A6A6"
    widths(ws, {"A": 22, "B": 78, "C": 26})
    header_row(ws, 1, ["Reviewer", "Client / SOA folder", "Note"])
    link_font = Font(name=FONT, size=10, color="FF0563C1", underline="single")
    out = 2
    for who, what, note in peer_history:
        ws.cell(row=out, column=1, value=who).font = BODY
        c = ws.cell(row=out, column=2, value=what)
        if isinstance(what, str) and what.startswith("http"):
            c.hyperlink = what
            c.font = link_font
        else:
            c.font = BODY
        ws.cell(row=out, column=3, value=note).font = BODY
        out += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{out - 1}"

    # Cancelled / RS -----------------------------------------------------
    ws = wb.create_sheet("RS-Cancelled History")
    ws.sheet_properties.tabColor = "FFA6A6A6"
    widths(ws, {"A": 34, "B": 14, "C": 18, "D": 11, "E": 14, "F": 14, "G": 40,
                "H": 40})
    header_row(ws, 1, ["Client", "Associate", "Status", "Division", "Adviser",
                       "SOA date", "Note", "Outcome"])
    out = 2
    for _, vals in cancelled:
        if all(v is None for v in vals):
            continue
        for i, v in enumerate(vals):
            cell = ws.cell(row=out, column=i + 1, value=v)
            cell.font = BODY
            if isinstance(v, dt.datetime):
                cell.number_format = DATE_FMT
        out += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{out - 1}"


def copy_legacy_2026(wb, src_path):
    """Keep a plain-values copy of the old 2026 tab as history."""
    src = openpyxl.load_workbook(src_path, data_only=True)["SOA Allocation 2026"]
    ws = wb.create_sheet("History 2026 (old layout)")
    ws.sheet_properties.tabColor = "FFA6A6A6"
    widths(ws, {"A": 7, "B": 8, "C": 14, "D": 38, "E": 14, "F": 14, "G": 11,
                "H": 14, "I": 14, "J": 46})
    header_row(ws, 1, ["Week", "Month", "Week label", "Client", "Associate",
                       "Client type", "Division", "Adviser", "SOA date",
                       "Notes"])
    out = 2
    for r in range(2, src.max_row + 1):
        vals = [src.cell(r, c).value for c in range(1, 11)]
        if all(v is None for v in vals):
            continue
        for i, v in enumerate(vals):
            cell = ws.cell(row=out, column=i + 1, value=v)
            cell.font = BODY
            if isinstance(v, dt.datetime):
                cell.number_format = DATE_FMT
        out += 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{out - 1}"


# --------------------------------------------------------------------------
# Defined names
# --------------------------------------------------------------------------
def add_names(wb, setup):
    rf, rl = setup["roster_first"], setup["roster_last"]
    lf, ll = setup["list_first"], setup["list_last"]
    cf, cl = setup["caps_first"], setup["caps_last"]
    hf, hl = setup["hol_first"], setup["hol_last"]
    vf, vl = setup["leave_first"], setup["leave_last"]
    la = LAST_ALLOC_ROW
    wl = N_WEEK_ROWS + 1

    names = {
        # Setup
        "RosterName": f"Setup!$A${rf}:$A${rl}",
        "RosterFull": f"Setup!$B${rf}:$B${rl}",
        "RosterActive": f"Setup!$D${rf}:$D${rl}",
        "RosterCap": f"Setup!$E${rf}:$E${rl}",
        "Associates": f"Setup!$K${rf}:$K${rl}",
        "Reviewers": f"Setup!$L${rf}:$L${rl}",
        "ReviewerCount": "Setup!$F$6",
        "TeamCapacity": "Setup!$F$7",
        "Advisers": f"Setup!$A${lf}:$A${ll}",
        "Divisions": f"Setup!$C${lf}:$C${ll}",
        "ClientTypes": f"Setup!$E${lf}:$E${ll}",
        "ProgressList": f"Setup!$G${lf}:$G${ll}",
        "ApptList": f"Setup!$I${lf}:$I${ll}",
        "ReviewStatusList": f"Setup!$K${lf}:$K${ll}",
        "CapDivision": f"Setup!$A${cf}:$A${cl}",
        "CapSOA": f"Setup!$B${cf}:$B${cl}",
        "HolDate": f"Setup!$A${hf}:$A${hl}",
        "HolName": f"Setup!$B${hf}:$B${hl}",
        "HolIdx": f"Setup!$E${hf}:$E${hl}",
        "LeaveWho": f"Setup!$A${vf}:$A${vl}",
        "LeaveFrom": f"Setup!$B${vf}:$B${vl}",
        "LeaveTo": f"Setup!$C${vf}:$C${vl}",
        "LeaveCap": f"Setup!$F${vf}:$F${vl}",
        # Weeks
        "WeekNo": f"Weeks!$A$2:$A${wl}",
        "WeekStart": f"Weeks!$B$2:$B${wl}",
        "WeekEnd": f"Weeks!$C$2:$C${wl}",
        "WeekMonth": f"Weeks!$D$2:$D${wl}",
        "WeekLabel": f"Weeks!$E$2:$E${wl}",
        "WeekHols": f"Weeks!$F$2:$F${wl}",
        "WeekNote": f"Weeks!$I$2:$I${wl}",
        "WeekAvail": f"Weeks!$M$2:$M${wl}",
        "WeekSpare": f"Weeks!$N$2:$N${wl}",
        "WeekDivHeads": "Weeks!$P$1:$V$1",
        "WeekDivBlock": f"Weeks!$P$2:$V${wl}",
        # Allocations
        "AllocWeek": f"Allocations!$A$2:$A${la}",
        "AllocClient": f"Allocations!$E$2:$E${la}",
        "AllocAssoc": f"Allocations!$F$2:$F${la}",
        "AllocType": f"Allocations!$H$2:$H${la}",
        "AllocDivision": f"Allocations!$I$2:$I${la}",
        "AllocAdviser": f"Allocations!$J$2:$J${la}",
        "AllocDate": f"Allocations!$K$2:$K${la}",
        "AllocAppt": f"Allocations!$L$2:$L${la}",
        "AllocProgress": f"Allocations!$M$2:$M${la}",
        "AllocReviewer": f"Allocations!$P$2:$P${la}",
        "AllocChecks": f"Allocations!$Q$2:$Q${la}",
        "AllocLink": f"Allocations!$S$2:$S${la}",
        "AllocRevStatus": f"Allocations!$T$2:$T${la}",
        "AllocReviewedOn": f"Allocations!$U$2:$U${la}",
        "AllocComments": f"Allocations!$V$2:$V${la}",
        "AllocKey": f"Allocations!$Z$2:$Z${la}",
        "AllocReviewKey": f"Allocations!$AB$2:$AB${la}",
    }
    for name, ref in names.items():
        wb.defined_names[name] = DefinedName(name, attr_text=ref)


# --------------------------------------------------------------------------
def main():
    global LAST_ALLOC_ROW
    src, out = sys.argv[1], sys.argv[2]
    if len(sys.argv) > 3:
        # Smaller register, used to verify the formulas quickly.
        LAST_ALLOC_ROW = int(sys.argv[3])
    raw, week_notes, leave, hist25, peer_hist, cancelled = parse_legacy(src)
    allocations = [normalise(a) for a in raw]

    # Leave markers on the old tab were one row per missing slot; collapse
    # them into one range per associate per week.
    leave_rows = []
    for (assoc, week), reason in leave.items():
        start = WEEK_ANCHOR + dt.timedelta(days=7 * (week - 1))
        leave_rows.append((assoc, start, start + dt.timedelta(days=4),
                           f"{reason} (from 2026 tab, week {week})"))
    leave_rows.sort(key=lambda x: (x[0], x[1]))
    # Merge consecutive weeks for the same person.
    merged = []
    for row in leave_rows:
        if merged and merged[-1][0] == row[0] and \
                merged[-1][2] + dt.timedelta(days=3) == row[1]:
            prev = merged[-1]
            merged[-1] = (prev[0], prev[1], row[2], prev[3])
        else:
            merged.append(row)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    setup = build_setup(wb, merged)
    build_weeks(wb, week_notes)
    build_allocations(wb, allocations)
    build_week_view(wb)
    build_dashboard(wb)
    build_peer_review(wb)
    build_history(wb, hist25, peer_hist, cancelled)
    copy_legacy_2026(wb, src)
    build_start_here(wb)
    add_names(wb, setup)

    wb.save(out)

    print(f"allocations migrated : {len(allocations)}")
    print(f"leave ranges         : {len(merged)}")
    print(f"week notes           : {len(week_notes)}")
    print(f"2025 history rows    : {len(hist25)}")
    print(f"peer review history  : {len(peer_hist)}")
    print(f"associates seen      : "
          f"{sorted(Counter(a['assoc'] for a in allocations))}")
    print(f"written              : {out}")


if __name__ == "__main__":
    main()
